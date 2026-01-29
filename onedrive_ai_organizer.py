import os
import re
import json
import time
import shutil
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import requests
from flask import Flask, request, redirect, url_for, render_template_string, jsonify

# Optional extractors
import pdfplumber
from docx import Document
from openpyxl import load_workbook
from PIL import Image

APP_TITLE = "OneDrive Downloads AI Organizer (Selective + Safe)"
STATE_FILE = "organizer_state.json"
ACTIONS_LOG = "organizer_actions_log.jsonl"

# ---- Configure your allowed folder taxonomy here ----
ALLOWED_FOLDERS = [
    "Finance/Invoices/2024",
    "Finance/Invoices/2025",
    "Finance/Taxes",
    "Finance/Contracts",
    "Work/IT",
    "Work/Docs",
    "Work/Clients",
    "Dev/Docs",
    "Dev/Assets",
    "Dev/Tools",
    "Design/AI",
    "Design/PSD",
    "Design/SVG",
    "Media/Images",
    "Media/Screenshots",
    "Installers",
    "_ToSort",
]

# ---- Ollama config ----
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
OLLAMA_MODEL = "llama3.1:8b"  # change if you use a different model

# ---- Extraction limits ----
MAX_TEXT_CHARS = 4000
MAX_FILES_SCAN = 5000

# ---- Safety defaults ----
DEFAULT_MODE = "move"   # "move" or "copy"
NEVER_OVERWRITE = True

# Minimal UI templates (single file)
BASE_HTML = r"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>{{title}}</title>
  <style>
    body { font-family: system-ui, Segoe UI, Arial; margin: 18px; }
    .row { display: flex; gap: 14px; flex-wrap: wrap; align-items: center; }
    .card { border: 1px solid #ddd; border-radius: 12px; padding: 14px; margin: 12px 0; }
    .muted { color: #666; font-size: 12px; }
    table { border-collapse: collapse; width: 100%; }
    th, td { border-bottom: 1px solid #eee; padding: 10px 8px; vertical-align: top; }
    th { text-align: left; position: sticky; top: 0; background: #fff; border-bottom: 1px solid #ddd; }
    input[type="text"] { width: 100%; padding: 8px; border-radius: 10px; border: 1px solid #ddd; }
    select { padding: 8px; border-radius: 10px; border: 1px solid #ddd; }
    .btn { display: inline-block; padding: 9px 12px; border-radius: 10px; border: 1px solid #ddd; background: #f7f7f7; cursor: pointer; text-decoration: none; color: #111; }
    .btn:hover { background: #efefef; }
    .danger { background: #ffecec; border-color: #ffb3b3; }
    .ok { background: #eaffea; border-color: #b6f0b6; }
    .tag { display: inline-block; padding: 2px 8px; border-radius: 999px; border: 1px solid #ddd; font-size: 12px; margin-right: 6px; }
    .tag.never { background: #fff0e6; border-color: #ffd1b3; }
    .tag.cand { background: #eaf3ff; border-color: #c6ddff; }
    .tag.done { background: #eaffea; border-color: #b6f0b6; }
    .small { font-size: 12px; }
    .mono { font-family: ui-monospace, Menlo, Consolas, monospace; font-size: 12px; }
  </style>
</head>
<body>
  <h2>{{title}}</h2>
  <div class="muted">Local only. Nothing happens until you approve. Logs written to <span class="mono">{{actions_log}}</span></div>
  <div style="height:10px"></div>
  {% block content %}{% endblock %}
</body>
</html>
"""

# ---------------- Helpers ----------------
def render_page(content_html: str, **ctx):
    page = BASE_HTML.replace("{% block content %}{% endblock %}", content_html)
    return render_template_string(page, **ctx)

# ---------------- State ------------------
def load_state() -> Dict[str, Any]:
    if os.path.exists(STATE_FILE):
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {
        "root": "",
        "mode": DEFAULT_MODE,
        "items": {},  # key: relpath -> data
        "allowed_folders": ALLOWED_FOLDERS,
    }

def save_state(state: Dict[str, Any]) -> None:
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, ensure_ascii=False)

def log_action(payload: Dict[str, Any]) -> None:
    payload["ts"] = datetime.now().isoformat(timespec="seconds")
    with open(ACTIONS_LOG, "a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")

# ---------------- Utilities ----------------
def sha256_file(path: Path, max_bytes: int = 2_000_000) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        remaining = max_bytes
        while remaining > 0:
            chunk = f.read(min(65536, remaining))
            if not chunk:
                break
            h.update(chunk)
            remaining -= len(chunk)
    return h.hexdigest()

def safe_filename(name: str) -> str:
    # keep it Windows-safe
    name = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    # avoid trailing dots/spaces in Windows
    name = name.rstrip(" .")
    return name[:180] if len(name) > 180 else name

def relpath_under(root: Path, full: Path) -> str:
    return str(full.relative_to(root)).replace("\\", "/")

def human_size(n: int) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if n < 1024:
            return f"{n:.0f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"

# ---------------- Content extraction ----------------
def extract_preview(path: Path) -> Dict[str, Any]:
    ext = path.suffix.lower()
    info: Dict[str, Any] = {"kind": "metadata", "text": "", "notes": ""}

    try:
        stat = path.stat()
        info["size"] = stat.st_size
        info["mtime"] = datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds")
    except Exception:
        pass

    # Quick metadata for risky/binary stuff
    if ext in [".exe", ".msi", ".dll"]:
        info["kind"] = "binary"
        try:
            info["hash"] = sha256_file(path)
        except Exception as e:
            info["notes"] = f"hash failed: {e}"
        return info

    if ext in [".zip", ".rar", ".7z"]:
        info["kind"] = "archive"
        # Don't unpack; just leave metadata. (We can add filename listing later if you want.)
        return info

    if ext in [".ai", ".psd"]:
        info["kind"] = "design"
        return info

    # Text-ish
    if ext in [".txt", ".md", ".log", ".csv"]:
        info["kind"] = "text"
        try:
            txt = path.read_text(encoding="utf-8", errors="ignore")
            info["text"] = txt[:MAX_TEXT_CHARS]
        except Exception as e:
            info["notes"] = f"read failed: {e}"
        return info

    if ext in [".json"]:
        info["kind"] = "json"
        try:
            raw = path.read_text(encoding="utf-8", errors="ignore")
            # summarize keys if possible
            try:
                obj = json.loads(raw)
                if isinstance(obj, dict):
                    keys = list(obj.keys())[:60]
                    info["text"] = f"JSON object keys: {keys}"
                elif isinstance(obj, list):
                    info["text"] = f"JSON array length: {len(obj)}; first item type: {type(obj[0]).__name__ if obj else 'empty'}"
                else:
                    info["text"] = f"JSON type: {type(obj).__name__}"
            except Exception:
                info["text"] = raw[:MAX_TEXT_CHARS]
        except Exception as e:
            info["notes"] = f"read failed: {e}"
        return info

    if ext in [".docx"]:
        info["kind"] = "docx"
        try:
            doc = Document(str(path))
            parts = []
            for p in doc.paragraphs[:80]:
                if p.text.strip():
                    parts.append(p.text.strip())
                if sum(len(x) for x in parts) > MAX_TEXT_CHARS:
                    break
            info["text"] = "\n".join(parts)[:MAX_TEXT_CHARS]
        except Exception as e:
            info["notes"] = f"docx parse failed: {e}"
        return info

    if ext in [".xlsx", ".xlsm"]:
        info["kind"] = "xlsx"
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
            sheets = wb.sheetnames[:20]
            preview_lines = [f"Sheets: {sheets}"]
            # grab a tiny header from first sheet
            if sheets:
                ws = wb[sheets[0]]
                header = []
                for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
                    if cell is None:
                        continue
                    header.append(str(cell)[:60])
                if header:
                    preview_lines.append(f"Header row: {header[:25]}")
            info["text"] = "\n".join(preview_lines)[:MAX_TEXT_CHARS]
        except Exception as e:
            info["notes"] = f"xlsx parse failed: {e}"
        return info

    if ext in [".pdf"]:
        info["kind"] = "pdf"
        try:
            text_parts = []
            with pdfplumber.open(str(path)) as pdf:
                for i, page in enumerate(pdf.pages[:3]):
                    t = page.extract_text() or ""
                    t = re.sub(r"\s+", " ", t).strip()
                    if t:
                        text_parts.append(t)
                    if sum(len(x) for x in text_parts) > MAX_TEXT_CHARS:
                        break
            info["text"] = "\n".join(text_parts)[:MAX_TEXT_CHARS]
        except Exception as e:
            info["notes"] = f"pdf parse failed: {e}"
        return info

    if ext in [".jpg", ".jpeg", ".png", ".webp"]:
        info["kind"] = "image"
        # No OCR by default (keeps it fast). We can add OCR toggle later.
        try:
            im = Image.open(path)
            info["notes"] = f"{im.width}x{im.height}"
        except Exception:
            pass
        return info

    if ext in [".svg"]:
        info["kind"] = "svg"
        try:
            raw = path.read_text(encoding="utf-8", errors="ignore")
            # pull <title> or first lines
            m = re.search(r"<title>(.*?)</title>", raw, flags=re.IGNORECASE | re.DOTALL)
            if m:
                info["text"] = f"SVG title: {re.sub(r'\\s+', ' ', m.group(1)).strip()}"[:MAX_TEXT_CHARS]
            else:
                info["text"] = raw[:800]
        except Exception as e:
            info["notes"] = f"svg read failed: {e}"
        return info

    # Default: metadata only
    return info

# ---------------- LLM suggestion ----------------
def ollama_suggest(filename: str, ext: str, preview: Dict[str, Any], allowed_folders: List[str]) -> Dict[str, Any]:
    # Prompt designed to return strict JSON
    content_hint = preview.get("text", "")
    kind = preview.get("kind", "metadata")

    system = (
        "You are a careful file organization assistant. "
        "You must return ONLY valid JSON (no markdown, no commentary). "
        "You propose a better filename and a destination folder from an allowed list. "
        "If uncertain, set confidence low and choose _ToSort."
    )

    user = {
        "original_filename": filename,
        "extension": ext,
        "kind": kind,
        "content_preview": content_hint[:MAX_TEXT_CHARS],
        "allowed_folders": allowed_folders,
        "rules": [
            "Destination folder MUST be one of allowed_folders exactly.",
            "Suggested filename must keep the same extension.",
            "Use clear names with dates if present (YYYY-MM-DD), otherwise omit date.",
            "Avoid overly long names; <= 80 characters before extension is ideal.",
            "If you can't infer, pick folder _ToSort and keep filename similar.",
            "Return JSON with keys: suggested_name, suggested_folder, confidence (0..1), reason."
        ]
    }

    payload = {
        "model": OLLAMA_MODEL,
        "prompt": system + "\n\nTASK:\n" + json.dumps(user, ensure_ascii=False),
        "stream": False,
        "options": {"temperature": 0.2}
    }

    r = requests.post(OLLAMA_URL, json=payload, timeout=600)
    r.raise_for_status()
    out = r.json().get("response", "").strip()

    # Try to parse JSON safely (model may add junk; strip around first/last braces)
    start = out.find("{")
    end = out.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return {
            "suggested_name": filename,
            "suggested_folder": "_ToSort",
            "confidence": 0.0,
            "reason": "Model did not return JSON."
        }
    jtxt = out[start:end+1]
    try:
        obj = json.loads(jtxt)
    except Exception:
        return {
            "suggested_name": filename,
            "suggested_folder": "_ToSort",
            "confidence": 0.0,
            "reason": "Could not parse model JSON."
        }

    # Normalize / enforce rules
    sug_name = safe_filename(obj.get("suggested_name", filename))
    if not sug_name.lower().endswith(ext.lower()):
        sug_name = safe_filename(Path(sug_name).stem + ext)

    folder = obj.get("suggested_folder", "_ToSort")
    if folder not in allowed_folders:
        folder = "_ToSort"

    conf = obj.get("confidence", 0.0)
    try:
        conf = float(conf)
    except Exception:
        conf = 0.0
    conf = max(0.0, min(1.0, conf))

    reason = str(obj.get("reason", "")).strip()[:300]

    return {
        "suggested_name": sug_name,
        "suggested_folder": folder,
        "confidence": conf,
        "reason": reason
    }

# ---------------- File ops ----------------
def apply_change(root: Path, rel: str, dest_folder: str, new_name: str, mode: str) -> Tuple[bool, str, str]:
    src = root / rel
    dest_dir = root / dest_folder
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / new_name

    if NEVER_OVERWRITE and dest.exists():
        return False, str(dest), "Destination exists (overwrite disabled)."

    try:
        if mode == "copy":
            shutil.copy2(src, dest)
        else:
            shutil.move(str(src), str(dest))
        return True, str(dest), ""
    except Exception as e:
        return False, str(dest), f"{type(e).__name__}: {e}"

# ---------------- Flask app ----------------
app = Flask(__name__)

@app.get("/")
def home():
    state = load_state()
    root = state.get("root", "")
    items = state.get("items", {})
    counts = {"total": len(items), "never": 0, "cand": 0, "done": 0}
    for v in items.values():
        st = v.get("status", "candidate")
        if st == "never":
            counts["never"] += 1
        elif st == "done":
            counts["done"] += 1
        else:
            counts["cand"] += 1

    html = """
      <div class="card">
        <div class="row">
          <form method="post" action="{{url_for('set_root')}}" style="flex:1; min-width:320px;">
            <div class="muted">Root folder to scan (your OneDrive Downloads path)</div>
            <div class="row">
              <input type="text" name="root" placeholder="C:/Users/You/OneDrive/Downloads" value="{{root}}" />
              <button class="btn" type="submit">Set</button>
            </div>
          </form>

          <form method="post" action="{{url_for('set_mode')}}" style="min-width:260px;">
            <div class="muted">Apply mode</div>
            <div class="row">
              <select name="mode">
                <option value="move" {% if mode=="move" %}selected{% endif %}>Move (default)</option>
                <option value="copy" {% if mode=="copy" %}selected{% endif %}>Copy (safer)</option>
              </select>
              <button class="btn" type="submit">Save</button>
            </div>
          </form>
        </div>

        <div style="height:8px"></div>
        <div class="row">
          <a class="btn" href="{{url_for('scan')}}">Scan folder</a>
          <a class="btn" href="{{url_for('review')}}">Review & tag (Never touch / Candidate)</a>
          <a class="btn" href="{{url_for('suggest')}}?limit=10">AI suggest (next 10)</a>
          <a class="btn" href="{{url_for('suggest')}}?limit=25">AI suggest (next 25)</a>
          <a class="btn ok" href="{{url_for('apply')}}">Apply approved changes</a>
        </div>
      </div>

      <div class="card">
        <div class="row">
          <div><span class="tag">Total: {{counts.total}}</span></div>
          <div><span class="tag never">Never touch: {{counts.never}}</span></div>
          <div><span class="tag cand">Candidates: {{counts.cand}}</span></div>
          <div><span class="tag done">Done: {{counts.done}}</span></div>
        </div>
        <div style="height:8px"></div>
        <div class="muted">
          Tip: Start with <b>Scan</b>, then in <b>Review</b> mark “Never touch”, then run <b>AI suggest</b>.
        </div>
      </div>
    """
    return render_page(html, base=BASE_HTML, title=APP_TITLE, actions_log=ACTIONS_LOG, root=root, mode=state.get("mode", DEFAULT_MODE), counts=counts)

@app.post("/set-root")
def set_root():
    state = load_state()
    state["root"] = request.form.get("root", "").strip().strip('"')
    save_state(state)
    return redirect(url_for("home"))

@app.post("/set-mode")
def set_mode():
    state = load_state()
    mode = request.form.get("mode", DEFAULT_MODE)
    if mode not in ["move", "copy"]:
        mode = DEFAULT_MODE
    state["mode"] = mode
    save_state(state)
    return redirect(url_for("home"))

@app.get("/scan")
def scan():
    state = load_state()
    root = state.get("root", "")
    if not root:
        return redirect(url_for("home"))

    rootp = Path(root)
    if not rootp.exists():
        return f"Root path does not exist: {root}", 400

    items = {}
    count = 0
    for p in rootp.rglob("*"):
        if count >= MAX_FILES_SCAN:
            break
        if p.is_file():
            rel = relpath_under(rootp, p)
            ext = p.suffix.lower()
            try:
                st = p.stat()
                items[rel] = {
                    "rel": rel,
                    "name": p.name,
                    "ext": ext,
                    "size": st.st_size,
                    "mtime": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
                    "status": "candidate",   # candidate | never | done
                    "approved": False,
                    "preview": None,         # filled lazily
                    "suggestion": None,      # filled later
                    "edited_name": "",
                    "edited_folder": "",
                }
                count += 1
            except Exception:
                continue

    state["items"] = items
    save_state(state)
    return redirect(url_for("review"))

@app.get("/review")
def review():
    state = load_state()
    root = state.get("root", "")
    items = state.get("items", {})
    q = request.args.get("q", "").strip().lower()
    filt = request.args.get("f", "all")  # all/candidate/never/done

    rows = []
    for rel, v in items.items():
        if filt != "all" and v.get("status") != filt:
            continue
        if q and (q not in rel.lower()):
            continue
        rows.append(v)

    rows.sort(key=lambda x: x["rel"].lower())

    html = """
      <div class="card">
        <div class="row">
          <a class="btn" href="{{url_for('home')}}">← Home</a>
          <form method="get" action="{{url_for('review')}}" style="flex:1; min-width:280px;">
            <input type="text" name="q" placeholder="Search path/name…" value="{{q}}" />
            <input type="hidden" name="f" value="{{filt}}" />
          </form>
          <div>
            <a class="btn" href="{{url_for('review', f='all', q=q)}}">All</a>
            <a class="btn" href="{{url_for('review', f='candidate', q=q)}}">Candidate</a>
            <a class="btn" href="{{url_for('review', f='never', q=q)}}">Never touch</a>
            <a class="btn" href="{{url_for('review', f='done', q=q)}}">Done</a>
          </div>
        </div>
        <div class="muted">Root: <span class="mono">{{root}}</span></div>
      </div>

      <div class="card">
        <form method="post" action="{{url_for('bulk_set_status')}}">
          <div class="row">
            <select name="new_status">
              <option value="candidate">Set selected → Candidate</option>
              <option value="never">Set selected → Never touch</option>
            </select>
            <button class="btn" type="submit">Apply to selected</button>
            <div class="muted small">Select files below, then mark them “Never touch”.</div>
          </div>

          <div style="height:10px"></div>
          <table>
            <tr>
              <th style="width:36px;"></th>
              <th>File</th>
              <th style="width:120px;">Size</th>
              <th style="width:170px;">Modified</th>
              <th style="width:130px;">Status</th>
              <th style="width:140px;">Preview</th>
            </tr>
            {% for r in rows %}
            <tr>
              <td><input type="checkbox" name="sel" value="{{r.rel}}"/></td>
              <td>
                <div class="mono">{{r.rel}}</div>
                <div class="muted">{{r.ext}}</div>
              </td>
              <td>{{human_size(r.size)}}</td>
              <td class="small">{{r.mtime}}</td>
              <td>
                {% if r.status=="never" %}
                  <span class="tag never">Never</span>
                {% elif r.status=="done" %}
                  <span class="tag done">Done</span>
                {% else %}
                  <span class="tag cand">Candidate</span>
                {% endif %}
              </td>
              <td>
                <a class="btn" href="{{url_for('preview', rel=r.rel)}}">Open</a>
              </td>
            </tr>
            {% endfor %}
          </table>
        </form>
      </div>
    """
    return render_page(
        html, base=BASE_HTML, title="Review & Tag", actions_log=ACTIONS_LOG,
        root=root, rows=rows, q=q, filt=filt, human_size=human_size
    )

@app.post("/bulk-set-status")
def bulk_set_status():
    state = load_state()
    items = state.get("items", {})
    new_status = request.form.get("new_status", "candidate")
    if new_status not in ["candidate", "never"]:
        new_status = "candidate"
    sels = request.form.getlist("sel")
    for rel in sels:
        if rel in items and items[rel].get("status") != "done":
            items[rel]["status"] = new_status
            items[rel]["approved"] = False
    state["items"] = items
    save_state(state)
    return redirect(url_for("review"))

@app.get("/preview")
def preview():
    state = load_state()
    root = state.get("root", "")
    rel = request.args.get("rel", "")
    items = state.get("items", {})
    if rel not in items:
        return redirect(url_for("review"))

    rootp = Path(root)
    fp = rootp / rel
    it = items[rel]

    if it.get("preview") is None:
        it["preview"] = extract_preview(fp)
        items[rel] = it
        state["items"] = items
        save_state(state)

    pv = it["preview"] or {}
    text = pv.get("text", "")
    notes = pv.get("notes", "")
    kind = pv.get("kind", "")

    html = """
      <div class="card">
        <div class="row">
          <a class="btn" href="{{url_for('review')}}">← Back</a>
          <div class="mono">{{rel}}</div>
          {% if status=="never" %}<span class="tag never">Never</span>{% endif %}
          {% if status=="candidate" %}<span class="tag cand">Candidate</span>{% endif %}
          {% if status=="done" %}<span class="tag done">Done</span>{% endif %}
        </div>
        <div class="muted">Kind: <b>{{kind}}</b> | Notes: {{notes}}</div>
      </div>

      <div class="card">
        <div class="muted">Extracted preview (for AI):</div>
        <pre style="white-space:pre-wrap; font-family: ui-monospace, Menlo, Consolas, monospace; font-size:12px;">{{text}}</pre>
      </div>
    """
    return render_page(
        html, base=BASE_HTML, title="Preview", actions_log=ACTIONS_LOG,
        rel=rel, kind=kind, notes=notes, text=text, status=it.get("status", "candidate")
    )

@app.get("/suggest")
def suggest():
    limit = int(request.args.get("limit", "10"))
    limit = max(1, min(50, limit))
    state = load_state()
    root = state.get("root", "")
    items = state.get("items", {})
    allowed = state.get("allowed_folders", ALLOWED_FOLDERS)

    rootp = Path(root)
    # Run suggestions for candidates that don't have a suggestion yet
    suggested = 0
    for rel, it in items.items():
        if it.get("status") != "candidate":
            continue
        if it.get("suggestion") is not None:
            continue

        fp = rootp / rel
        if it.get("preview") is None:
            it["preview"] = extract_preview(fp)

        sug = ollama_suggest(it["name"], it["ext"], it["preview"], allowed)
        it["suggestion"] = sug
        it["edited_name"] = sug["suggested_name"]
        it["edited_folder"] = sug["suggested_folder"]
        it["approved"] = sug.get("confidence", 0.0) >= 0.75  # auto-check only if high confidence
        items[rel] = it
        suggested += 1

        if suggested >= limit:
            break

        # tiny pause so UI doesn't feel frozen on slower PCs
        time.sleep(0.05)

    state["items"] = items
    save_state(state)
    return redirect(url_for("proposals"))

@app.get("/proposals")
def proposals():
    state = load_state()
    items = state.get("items", {})
    allowed = state.get("allowed_folders", ALLOWED_FOLDERS)

    rows = []
    for it in items.values():
        if it.get("status") == "candidate" and it.get("suggestion") is not None:
            rows.append(it)
    rows.sort(key=lambda x: (-(x["suggestion"].get("confidence", 0.0)), x["rel"].lower()))

    html = """
      <div class="card">
        <div class="row">
          <a class="btn" href="{{url_for('home')}}">← Home</a>
          <a class="btn" href="{{url_for('review')}}">Review tags</a>
          <a class="btn" href="{{url_for('suggest')}}">Run AI (remaining)</a>
          <a class="btn ok" href="{{url_for('apply')}}">Apply approved</a>
        </div>
        <div class="muted">Auto-approved if confidence ≥ 0.75 (you can change). Nothing applied yet.</div>
      </div>

      <div class="card">
        <form method="post" action="{{url_for('update_proposals')}}">
          <table>
            <tr>
              <th style="width:72px;">Approve</th>
              <th>Original</th>
              <th>New folder</th>
              <th>New name</th>
              <th style="width:90px;">Conf.</th>
              <th>Reason</th>
            </tr>
            {% for r in rows %}
            <tr>
              <td>
                <input type="checkbox" name="appr" value="{{r.rel}}" {% if r.approved %}checked{% endif %}/>
              </td>
              <td>
                <div class="mono">{{r.rel}}</div>
                <div class="muted small"><a href="{{url_for('preview', rel=r.rel)}}">preview</a></div>
              </td>
              <td>
                <select name="folder__{{r.rel}}">
                  {% for f in allowed %}
                    <option value="{{f}}" {% if r.edited_folder==f %}selected{% endif %}>{{f}}</option>
                  {% endfor %}
                </select>
              </td>
              <td>
                <input type="text" name="name__{{r.rel}}" value="{{r.edited_name}}"/>
              </td>
              <td class="mono">{{"%.2f"|format(r.suggestion.confidence)}}</td>
              <td class="small">{{r.suggestion.reason}}</td>
            </tr>
            {% endfor %}
          </table>

          <div style="height:10px"></div>
          <div class="row">
            <button class="btn" type="submit">Save edits</button>
            <a class="btn ok" href="{{url_for('apply')}}">Apply approved</a>
          </div>
        </form>
      </div>
    """
    # Jinja can't access dot keys reliably, so map suggestion keys
    for r in rows:
        r["suggestion"] = r["suggestion"] or {}
        r["suggestion"]["confidence"] = float(r["suggestion"].get("confidence", 0.0))
        r["suggestion"]["reason"] = r["suggestion"].get("reason", "")
    return render_page(
        html, base=BASE_HTML, title="AI Proposals", actions_log=ACTIONS_LOG,
        rows=rows, allowed=allowed
    )

@app.post("/update-proposals")
def update_proposals():
    state = load_state()
    items = state.get("items", {})
    approved_list = set(request.form.getlist("appr"))

    # Update edited fields
    for rel, it in items.items():
        if it.get("status") != "candidate" or it.get("suggestion") is None:
            continue
        it["approved"] = (rel in approved_list)
        new_folder = request.form.get(f"folder__{rel}", it.get("edited_folder", "_ToSort"))
        new_name = request.form.get(f"name__{rel}", it.get("edited_name", it.get("name", "")))
        it["edited_folder"] = new_folder
        it["edited_name"] = safe_filename(new_name)
        items[rel] = it

    state["items"] = items
    save_state(state)
    return redirect(url_for("proposals"))

@app.get("/apply")
def apply():
    state = load_state()
    root = state.get("root", "")
    mode = state.get("mode", DEFAULT_MODE)
    items = state.get("items", {})
    rootp = Path(root)

    results = []
    for rel, it in list(items.items()):
        if it.get("status") != "candidate":
            continue
        if not it.get("approved"):
            continue
        dest_folder = it.get("edited_folder") or "_ToSort"
        new_name = it.get("edited_name") or it.get("name")

        ok, dest, err = apply_change(rootp, rel, dest_folder, new_name, mode=mode)
        results.append({"rel": rel, "dest": dest, "ok": ok, "err": err, "mode": mode})

        log_action({
            "action": "apply",
            "mode": mode,
            "src": str(rootp / rel),
            "dest": dest,
            "ok": ok,
            "error": err
        })

        if ok:
            # mark done, and remove from items (since file moved); keep a done record
            it["status"] = "done"
            it["approved"] = False
            it["done_dest"] = dest
            items[rel] = it

    state["items"] = items
    save_state(state)

    html = """
      <div class="card">
        <div class="row">
          <a class="btn" href="{{url_for('home')}}">← Home</a>
          <a class="btn" href="{{url_for('proposals')}}">Back to proposals</a>
        </div>
        <div class="muted">Applied {{results|length}} approved items (mode: <b>{{mode}}</b>).</div>
      </div>

      <div class="card">
        <table>
          <tr>
            <th>Result</th><th>Original</th><th>Destination</th><th>Error</th>
          </tr>
          {% for r in results %}
          <tr>
            <td>{% if r.ok %}<span class="tag done">OK</span>{% else %}<span class="tag never">FAIL</span>{% endif %}</td>
            <td class="mono">{{r.rel}}</td>
            <td class="mono">{{r.dest}}</td>
            <td class="small">{{r.err}}</td>
          </tr>
          {% endfor %}
        </table>
      </div>
    """
    return render_page(
        html, base=BASE_HTML, title="Apply Results", actions_log=ACTIONS_LOG,
        results=results, mode=mode
    )

if __name__ == "__main__":
    # Save initial state if missing
    if not os.path.exists(STATE_FILE):
        save_state(load_state())

    print(f"{APP_TITLE}")
    print("Open: http://127.0.0.1:5000")
    app.run(host="127.0.0.1", port=5000, debug=False)
